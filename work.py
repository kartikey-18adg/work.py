import os
import pandas as pd
import numpy as np
import matplotlib

matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

class TraderSentimentAnalysis:
    def __init__(self, trader_file, fear_greed_file, output_file):
        self.trader_file = trader_file
        self.fear_greed_file = fear_greed_file
        self.output_file = output_file
        self.trader_data = self.load_trader_data()
        self.fear_greed_data = self.load_fear_greed_data()
        self.merged_data = pd.merge(
            self.trader_data, 
            self.fear_greed_data, 
            on="Date", 
            how="inner"
        )

    def load_trader_data(self):
        df = pd.read_csv(self.trader_file)
        column_mapping = {
            "Account": "account",
            "Coin": "symbol",
            "Execution Price": "execution_price",
            "Size Tokens": "size",
            "Side": "side",
            "Timestamp IST": "time",   
            "Start Position": "start_position",
            "Direction": "event",
            "Closed PnL": "closedPnL"
        }
        df = df.rename(columns=column_mapping)
        if "time" in df.columns:
            df["time"] = pd.to_datetime(df["time"], dayfirst=True, errors="coerce")
        else:
            raise KeyError("No valid time column found in trader data")
        df["Date"] = df["time"].dt.date
        if "leverage" not in df.columns:
            df["leverage"] = np.nan
        return df
    def load_fear_greed_data(self):
        df = pd.read_csv(self.fear_greed_file)
        df = df.rename(columns={
            'date': 'Date',
            'classification': 'Classification'
        })
        df['Date'] = pd.to_datetime(df['Date']).dt.date
        return df
    def calculate_trader_metrics(self):
        trader_metrics = (
            self.merged_data
            .groupby(["account", "Date", "Classification"])
            .agg(
                total_pnl=("closedPnL", "sum"),
                avg_leverage=("leverage", "mean"),
                trade_count=("account", "count"),
                avg_size=("size", "mean")
            )
            .reset_index()
        )
        return trader_metrics
    def analyze_sentiment_distribution(self):
        sentiment_counts = self.fear_greed_data['Classification'].value_counts()
        plt.figure(figsize=(6, 6))
        plt.pie(sentiment_counts, labels=sentiment_counts.index, autopct='%1.1f%%')
        plt.title("Fear/Greed Distribution")
        plt.savefig("sentiment_distribution.png")
        plt.close()
    def analyze_performance_by_sentiment(self):
        sentiment_perf = (
            self.merged_data.groupby("Classification")["closedPnL"].mean()
        )
        plt.figure(figsize=(8, 6))
        colors = sns.color_palette("Set2", n_colors=len(sentiment_perf))
        sentiment_perf.plot(kind="bar", color=colors)
        plt.title("Avg Trader PnL by Sentiment")
        plt.ylabel("Average PnL")
        plt.savefig("performance_by_sentiment.png")
        plt.close()
    def analyze_top_performers(self):
        top_traders = (
            self.merged_data.groupby("account")["closedPnL"]
            .sum()
            .sort_values(ascending=False)
            .head(10)
        )
        plt.figure(figsize=(10, 6))
        top_traders.plot(kind="bar")
        plt.title("Top 10 Traders by PnL")
        plt.ylabel("Total PnL")
        plt.savefig("top_traders.png")
        plt.close()
    def sentiment_transition_analysis(self):
        self.fear_greed_data["Prev_Sentiment"] = self.fear_greed_data["Classification"].shift(1)
        self.fear_greed_data["Transition"] = (
            self.fear_greed_data["Prev_Sentiment"] + " → " + self.fear_greed_data["Classification"]
        )
        transition_counts = self.fear_greed_data["Transition"].value_counts()
        plt.figure(figsize=(10, 6))
        transition_counts.plot(kind="bar", color="skyblue")
        plt.title("Sentiment Transitions Over Time")
        plt.savefig("sentiment_transitions.png")
        plt.close()
    def generate_trading_insights(self):
        insights = []
        sentiment_perf = self.merged_data.groupby("Classification")["closedPnL"].mean()
        if sentiment_perf.get("Fear", 0) < 0:
            insights.append("Traders tend to lose more during Fear sentiment.")
        if sentiment_perf.get("Greed", 0) > 0:
            insights.append("Traders perform better during Greed sentiment.")
        avg_leverage = self.merged_data.groupby("Classification")["leverage"].mean()
        insights.append(f"Avg leverage during Fear: {avg_leverage.get('Fear', 0):.2f}")
        insights.append(f"Avg leverage during Greed: {avg_leverage.get('Greed', 0):.2f}")
        return insights
    def export_to_excel(self):
        trader_metrics = self.calculate_trader_metrics()
        with pd.ExcelWriter(self.output_file, engine="openpyxl") as writer:
            trader_metrics.to_excel(writer, sheet_name="Trader Metrics", index=False)
            self.merged_data.to_excel(writer, sheet_name="Merged Data", index=False)
        self.style_excel_workbook()
    def style_excel_workbook(self):
        wb = openpyxl.load_workbook(self.output_file)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            if ws.max_row > 0:
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
        wb.save(self.output_file)
    def run_all_analyses(self):
        self.analyze_sentiment_distribution()
        self.analyze_performance_by_sentiment()
        self.analyze_top_performers()
        self.sentiment_transition_analysis()
        insights = self.generate_trading_insights()
        print("\nGenerated Insights:")
        for i in insights:
            print("-", i)
if __name__ == "__main__":
    trader_file = r"D:\k.cert\primetrade\historical_data.csv"
    fear_greed_file = r"D:\k.cert\primetrade\fear_greed_index.csv"
    output_file = r"D:\k.cert\primetrade\trader_sentiment_analysis.xlsx"
    analysis = TraderSentimentAnalysis(trader_file, fear_greed_file, output_file)
    analysis.run_all_analyses()
    analysis.export_to_excel()
    print("\nAnalysis complete. Results exported to:", output_file)